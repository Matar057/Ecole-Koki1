import io
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Avg
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import xlsxwriter
from .models import Exam, ExamSubject, Mark
from .forms import ExamForm, ExamSubjectForm, MarkForm
from users.decorators import role_required
from students.models import Student
from academics.models import Class


@login_required
def exam_list(request):
    exams = Exam.objects.select_related('academic_year').all()
    class_id = request.GET.get('classe')
    selected_class = None
    if class_id:
        selected_class = get_object_or_404(Class, pk=class_id)
        exams = exams.filter(subjects__class_obj=selected_class).distinct()
    return render(request, 'exams/exam_list.html', {
        'exams': exams,
        'selected_class': selected_class,
    })


@role_required('admin', 'teacher')
def exam_create(request):
    if request.method == 'POST':
        form = ExamForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Examen créé avec succès.')
            return redirect('exams:exam_list')
    else:
        form = ExamForm()
    return render(request, 'exams/exam_form.html', {'form': form, 'title': 'Créer un examen'})


@role_required('admin', 'teacher')
def exam_edit(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    if request.method == 'POST':
        form = ExamForm(request.POST, instance=exam)
        if form.is_valid():
            form.save()
            messages.success(request, 'Examen mis à jour.')
            return redirect('exams:exam_list')
    else:
        form = ExamForm(instance=exam)
    return render(request, 'exams/exam_form.html', {'form': form, 'title': "Modifier l'examen"})


@login_required
def exam_detail(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    exam_subjects = exam.subjects.select_related('subject', 'class_obj').all()
    return render(request, 'exams/exam_detail.html', {'exam': exam, 'exam_subjects': exam_subjects})


@role_required('admin', 'teacher')
def exam_subject_add(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    if request.method == 'POST':
        form = ExamSubjectForm(request.POST)
        if form.is_valid():
            es = form.save(commit=False)
            es.exam = exam
            es.save()
            messages.success(request, 'Matière d\'examen ajoutée.')
            return redirect('exams:exam_detail', pk=pk)
    else:
        form = ExamSubjectForm(initial={'exam': exam})
    return render(request, 'exams/exam_subject_form.html', {'form': form, 'exam': exam})


@login_required
def marks_list(request, exam_subject_id):
    exam_subject = get_object_or_404(ExamSubject, pk=exam_subject_id)
    marks = Mark.objects.filter(
        exam_subject=exam_subject
    ).select_related('student__user').order_by('-marks_obtained')
    total_students = Student.objects.filter(
        class_enrolled=exam_subject.class_obj, is_active=True
    ).count()
    marks_count = marks.count()
    avg_marks = marks.aggregate(Avg('marks_obtained'))['marks_obtained__avg'] or 0
    avg_note10 = round((avg_marks / exam_subject.max_marks) * 10, 2) if exam_subject.max_marks else 0
    return render(request, 'exams/marks_list.html', {
        'exam_subject': exam_subject,
        'marks': marks,
        'total_students': total_students,
        'marks_count': marks_count,
        'avg_marks': round(avg_marks, 2),
        'avg_note10': avg_note10,
    })


@role_required('admin', 'teacher')
def marks_entry(request, exam_subject_id):
    exam_subject = get_object_or_404(ExamSubject, pk=exam_subject_id)
    students = Student.objects.filter(class_enrolled=exam_subject.class_obj, is_active=True)
    if request.method == 'POST':
        for student in students:
            marks = request.POST.get(f'marks_{student.pk}')
            if marks:
                Mark.objects.update_or_create(
                    exam_subject=exam_subject, student=student,
                    defaults={'marks_obtained': float(marks), 'entered_by': request.user}
                )
        messages.success(request, 'Notes enregistrées avec succès.')
        return redirect('exams:marks_entry', exam_subject_id=exam_subject_id)
    marks = {m.student.pk: m.marks_obtained for m in exam_subject.marks.all()}
    student_marks = []
    for student in students:
        student_marks.append({
            'student': student,
            'marks': marks.get(student.pk, ''),
        })
    return render(request, 'exams/marks_entry.html', {
        'exam_subject': exam_subject,
        'student_marks': student_marks,
    })


@role_required('admin', 'teacher')
def mark_edit(request, pk):
    mark = get_object_or_404(Mark, pk=pk)
    if request.method == 'POST':
        form = MarkForm(request.POST, instance=mark)
        if form.is_valid():
            form.save()
            messages.success(request, 'Note modifiée avec succès.')
            return redirect('exams:marks_list', exam_subject_id=mark.exam_subject.pk)
    else:
        form = MarkForm(instance=mark)
    return render(request, 'exams/mark_form.html', {
        'form': form,
        'mark': mark,
        'title': 'Modifier la note',
    })


@role_required('admin', 'teacher')
def mark_delete(request, pk):
    mark = get_object_or_404(Mark, pk=pk)
    exam_subject_id = mark.exam_subject.pk
    if request.method == 'POST':
        mark.delete()
        messages.success(request, 'Note supprimée.')
        return redirect('exams:marks_list', exam_subject_id=exam_subject_id)
    return render(request, 'exams/mark_confirm_delete.html', {'mark': mark})


@login_required
def class_results(request, class_id, exam_id):
    class_obj = get_object_or_404(Class, pk=class_id)
    exam = get_object_or_404(Exam, pk=exam_id)
    exam_subjects = ExamSubject.objects.filter(exam=exam, class_obj=class_obj).select_related('subject')
    students = Student.objects.filter(class_enrolled=class_obj, is_active=True).select_related('user').order_by('user__last_name', 'user__first_name')
    marks = Mark.objects.filter(
        exam_subject__exam=exam, student__class_enrolled=class_obj
    ).select_related('exam_subject__subject', 'student__user')

    marks_by_student = {}
    for m in marks:
        sid = m.student.pk
        if sid not in marks_by_student:
            marks_by_student[sid] = {}
        marks_by_student[sid][m.exam_subject.pk] = m

    student_results = []
    for student in students:
        notes_sur_10 = []
        subject_data = []
        all_entered = True
        for es in exam_subjects:
            m = marks_by_student.get(student.pk, {}).get(es.pk)
            if m:
                note = float(m.marks_obtained) / es.max_marks * 10 if es.max_marks else 0
                notes_sur_10.append(note)
                subject_data.append({
                    'mark_pk': m.pk,
                    'marks_obtained': m.marks_obtained,
                    'max_marks': es.max_marks,
                    'note_sur_10': round(note, 2),
                    'grade': m.grade,
                })
            else:
                subject_data.append(None)
                all_entered = False
        moyenne = round(sum(notes_sur_10) / len(notes_sur_10), 1) if notes_sur_10 else 0
        student_results.append({
            'student': student,
            'subjects': subject_data,
            'moyenne': moyenne,
            'all_entered': all_entered,
        })

    student_results.sort(key=lambda r: r['moyenne'], reverse=True)
    for i, r in enumerate(student_results, 1):
        r['rank'] = i

    entered_results = [r for r in student_results if r['all_entered']]
    class_avg = round(sum(r['moyenne'] for r in entered_results) / len(entered_results), 1) if entered_results else 0

    return render(request, 'exams/class_results.html', {
        'class_obj': class_obj,
        'exam': exam,
        'exam_subjects': exam_subjects,
        'student_results': student_results,
        'class_avg': round(class_avg, 1),
    })


@login_required
def report_card(request, student_id, exam_id):
    student = get_object_or_404(Student, pk=student_id)
    exam = get_object_or_404(Exam, pk=exam_id)
    marks = Mark.objects.filter(
        exam_subject__exam=exam, student=student
    ).select_related('exam_subject__subject')

    class_obj = student.class_enrolled
    exam_subjects = ExamSubject.objects.filter(exam=exam, class_obj=class_obj).select_related('subject')
    classmates = Student.objects.filter(class_enrolled=class_obj, is_active=True).select_related('user')
    all_marks = Mark.objects.filter(exam_subject__exam=exam, student__class_enrolled=class_obj)

    notes = []
    for m in marks:
        note10 = float(m.marks_obtained) / m.exam_subject.max_marks * 10 if m.exam_subject.max_marks else 0
        notes.append(note10)
    moyenne = round(sum(notes) / len(notes), 2) if notes else 0
    max_marks = sum(m.exam_subject.max_marks for m in marks)

    class_avg_per_subject = {}
    for es in exam_subjects:
        subj_marks = [float(m.marks_obtained) / es.max_marks * 10 for m in all_marks if m.exam_subject_id == es.pk]
        class_avg_per_subject[es.pk] = round(sum(subj_marks) / len(subj_marks), 1) if subj_marks else 0

    for m in marks:
        m.note10 = round(float(m.marks_obtained) / m.exam_subject.max_marks * 10, 2) if m.exam_subject.max_marks else 0
        m.class_avg = class_avg_per_subject.get(m.exam_subject_id, 0)

    student_totals = []
    for c in classmates:
        c_marks = [m for m in all_marks if m.student_id == c.pk]
        c_notes = []
        for cm in c_marks:
            cn = float(cm.marks_obtained) / cm.exam_subject.max_marks * 10 if cm.exam_subject.max_marks else 0
            c_notes.append(cn)
        c_moy = round(sum(c_notes) / len(c_notes), 2) if c_notes else 0
        student_totals.append({'student_id': c.pk, 'moy': c_moy})

    student_totals.sort(key=lambda x: x['moy'], reverse=True)
    rank = 1
    for i, st in enumerate(student_totals):
        if st['student_id'] == student.pk:
            rank = i + 1
            break

    class_moys = [st['moy'] for st in student_totals if st['moy'] > 0]
    class_avg = round(sum(class_moys) / len(class_moys), 1) if class_moys else 0

    passe = moyenne >= 5
    next_class = Class.objects.filter(order=class_obj.order + 1, academic_year=class_obj.academic_year).first() if class_obj else None

    return render(request, 'exams/report_card.html', {
        'student': student, 'exam': exam, 'marks': marks,
        'moyenne': moyenne, 'max_marks': max_marks,
        'rank': rank, 'total_students': len(classmates),
        'class_avg': class_avg,
        'passe': passe,
        'next_class': next_class,
    })

@role_required('admin', 'teacher')
def marks_export(request, exam_subject_id):
    exam_subject = get_object_or_404(ExamSubject, pk=exam_subject_id)
    students = Student.objects.filter(class_enrolled=exam_subject.class_obj, is_active=True).order_by('user__last_name')
    marks = {m.student.pk: m.marks_obtained for m in exam_subject.marks.all()}

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('Notes')

    header_format = workbook.add_format({'bold': True, 'bg_color': '#4472C4', 'font_color': 'white', 'border': 1})
    cell_format = workbook.add_format({'border': 1, 'num_format': '0.00'})

    worksheet.write(0, 0, 'ID Étudiant', header_format)
    worksheet.write(0, 1, 'Nom complet', header_format)
    worksheet.write(0, 2, f'Note /{exam_subject.max_marks}', header_format)

    for i, student in enumerate(students, 1):
        worksheet.write(i, 0, student.student_id, cell_format)
        worksheet.write(i, 1, student.user.get_full_name(), cell_format)
        note = marks.get(student.pk, '')
        worksheet.write(i, 2, float(note) if note else '', cell_format)

    worksheet.set_column(0, 0, 18)
    worksheet.set_column(1, 1, 35)
    worksheet.set_column(2, 2, 15)

    workbook.close()
    output.seek(0)

    filename = f"notes_{exam_subject.subject.name}_{exam_subject.class_obj}.xlsx".replace(' ', '_')
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@role_required('admin', 'teacher')
def marks_import(request, exam_subject_id):
    exam_subject = get_object_or_404(ExamSubject, pk=exam_subject_id)

    if request.method == 'POST' and request.FILES.get('file'):
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(request.FILES['file'])
            sheet = workbook.active

            students_map = {s.student_id: s for s in Student.objects.filter(
                class_enrolled=exam_subject.class_obj, is_active=True
            )}

            imported = 0
            errors = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                student_id, _, note = row
                if student_id is None:
                    continue
                student = students_map.get(str(student_id).strip())
                if not student:
                    errors.append(f"Étudiant introuvable : {student_id}")
                    continue
                if note is None or note == '':
                    continue
                try:
                    note_val = float(note)
                    if note_val < 0 or note_val > exam_subject.max_marks:
                        errors.append(f"Note hors limite pour {student_id} : {note_val}")
                        continue
                    Mark.objects.update_or_create(
                        exam_subject=exam_subject, student=student,
                        defaults={'marks_obtained': note_val, 'entered_by': request.user}
                    )
                    imported += 1
                except ValueError:
                    errors.append(f"Note invalide pour {student_id} : {note}")

            if imported:
                messages.success(request, f'{imported} notes importées avec succès.')
            if errors:
                messages.warning(request, 'Erreurs : ' + '; '.join(errors[:5]))
            if not imported and not errors:
                messages.info(request, 'Aucune note à importer.')

        except Exception as e:
            messages.error(request, f'Erreur lors de l\'import : {e}')
        return redirect('exams:marks_entry', exam_subject_id=exam_subject_id)

    return render(request, 'exams/marks_import.html', {'exam_subject': exam_subject})


@role_required('admin')
def exam_delete(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    if request.method == 'POST':
        exam.delete()
        messages.success(request, 'Examen supprimé avec succès.')
        return redirect('exams:exam_list')
    return render(request, 'exams/exam_confirm_delete.html', {'exam': exam})


@role_required('admin', 'teacher')
def report_card_pdf(request, student_id, exam_id):
    student = get_object_or_404(Student, pk=student_id)
    exam = get_object_or_404(Exam, pk=exam_id)
    marks = Mark.objects.filter(
        exam_subject__exam=exam, student=student
    ).select_related('exam_subject__subject')
    total_marks = sum(m.marks_obtained for m in marks)
    max_marks = sum(m.exam_subject.max_marks for m in marks)
    percentage = (total_marks / max_marks * 100) if max_marks > 0 else 0
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=report_card_{student.student_id}_{exam.name}.pdf'
    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', fontSize=18, alignment=1, spaceAfter=10)
    elements = []
    elements.append(Paragraph('Système de gestion scolaire', title_style))
    elements.append(Paragraph(f'Bulletin de notes - {exam.name}', styles['h2']))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f'Étudiant: {student.user.get_full_name()}', styles['Normal']))
    elements.append(Paragraph(f'ID: {student.student_id}', styles['Normal']))
    elements.append(Paragraph(f'Classe: {student.class_enrolled}', styles['Normal']))
    elements.append(Spacer(1, 20))
    data = [['Matière', 'Note max', 'Obtenue', 'Grade']]
    for m in marks:
        data.append([
            m.exam_subject.subject.name,
            str(m.exam_subject.max_marks),
            str(m.marks_obtained),
            m.grade,
        ])
    data.append(['Total', str(max_marks), str(total_marks), f'{percentage:.1f}%'])
    table = Table(data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -2), 11),
    ]))
    elements.append(table)
    doc.build(elements)
    return response
